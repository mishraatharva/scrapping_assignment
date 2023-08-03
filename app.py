from flask import Flask
from flask import request, render_template
import requests
import os
from bs4 import BeautifulSoup
# from urllib.request import urlopen
import openpyxl as xl
from openpyxl import Workbook,  load_workbook

app = Flask(__name__)

file = "user_review.xlsx"
def save_to_excel(rating, review_rating):
    if not os.path.exists(file):
        workbook = Workbook()
        workbook.save(file)

    wBook = load_workbook(file)
    sheet = wBook.active
    data = [rating, review_rating]
    sheet.append(data)
    wBook.save(file)

def perform_scraping(link):
    response = requests.get(link)
    soup = BeautifulSoup(response.content, "html.parser")
    all_div = soup.find_all("div", class_="_1AtVbE col-12-12")
    del all_div[0]
    all_reviews = []
    for review in range(0, len(all_div)):
        rating_elem = all_div[review].find("div", class_="_3LWZlK _1BLPMq")
        review_title_elem = all_div[review].find("p", class_="_2-N8zT")

        if rating_elem is not None and review_title_elem is not None:
            rating = rating_elem.text
            review_title = review_title_elem.text
            save_to_excel(rating, review_title)
            all_reviews.append((rating, review_title))
    return all_reviews

@app.route('/index')
def index():
    return render_template('index.html')

all_reviews = []
@app.route("/savedata", methods = ['POST'])
def save_data():
    if request.method == 'POST':
        try:
            for page_number in range(1,3):
                link = f"https://www.flipkart.com/prowl-tiger-shroff-push-up-board-upper-body-workout-push-up-bar/product-reviews/itm0487671f4df34?pid=BAAGM82GUHTQ3KFZ&lid=LSTBAAGM82GUHTQ3KFZVCWDQ9&marketplace=FLIPKART&page={page_number}"
                all_reviews.extend(perform_scraping(link))
                return render_template("confirmation.html", result="Scraping Done, Click below link to Check!")
        except:
            return "<h3>some thing went wrong</h3>"

data_dict = {}
@app.route('/showreviews', methods = ['GET'])
def show_scraped_data():
    if request.method == 'GET':
        if not os.path.exists(file):
            return "<h3>data not exist</h3>"
        else:
            wBook = xl.load_workbook(file)
            sheet = wBook.active
            nrows= sheet.max_row
            print(nrows)
            ncolumns = sheet.max_column
            print(ncolumns)
            row = 1
            for i in range(1,nrows+1):
                for j in range(1,ncolumns+1,2):
                    cell_obj = sheet.cell(row=i,column=j)
                    data_dict[sheet.cell(row=i,column=j+1).value] = sheet.cell(row=i,column=j).value
    print(len(data_dict))
    return render_template("reviews.html", result= data_dict)

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8000,debug = True)